import dxpy
import re
from datetime import datetime, date
from dateutil import parser as date_parser
from openpyxl import load_workbook
import pandas as pd
import numpy as np
import os
import requests
import json
import uuid
import time

def get_folder(filename: str) -> str:
    """
    get the folder of input file
    Parameters:
    ----------
      str for input filename

    Return:
      str for folder name
    """
    folder = os.path.basename(os.path.normpath(os.path.dirname(filename)))
    print(folder)
    return folder


def get_summary_fields(
    filename: str, config_variable: dict, unusual_sample_name: bool
):  # -> tuple[pd.DataFrame, str]
    """
    Extract data from summary sheet of variant workbook

    Parameters
    ----------
      variant workbook file name
      dict from config file
      boolean for unusual_sample_name

    Returns
    -------
      data frame from summary sheet
      str for error message
    """
    workbook = load_workbook(filename)
    sampleID = workbook["summary"]["B1"].value
    clinical_indication = workbook["summary"]["F1"].value
    if ";" in clinical_indication:
        split_CI = clinical_indication.split(";")
        indication = []
        Rcode = []
        for each in split_CI:
            remove_R = each.split("_")[1]
            indication.append(remove_R)
            Rcode.append(each.split("_")[0])
        new_CI = ";".join(indication)
        combined_Rcode = ";".join(Rcode)
    else:
        new_CI = clinical_indication.split("_")[1]
        combined_Rcode = clinical_indication.split("_")[0]

    panel = workbook["summary"]["F2"].value
    date_evaluated = workbook["summary"]["G22"].value
    instrument, sample, batch, testcode, x, probeset = sampleID.split("-")
    ref_genome = "not_defined"
    for cell in workbook["summary"]["A"]:
        if cell.value == "Reference:":
            ref_genome = workbook["summary"][f"B{cell.row}"].value

    # checking sample naming
    error_msg = None
    if not unusual_sample_name:
        check_sample_name(
            instrument, sample, batch, testcode, probeset
        )
    d = {
        "Instrument ID": instrument,
        "Specimen ID": sample,
        "Batch ID": batch,
        "Test code": testcode,
        "Probeset ID": probeset,
        "Rcode": combined_Rcode,
        "Preferred condition name": new_CI,
        "Panel": panel,
        "Ref_genome": ref_genome,
        "Date last evaluated": date_evaluated,
    }
    df_summary = pd.DataFrame([d])

    # If no date last evaluated, use today's date
    df_summary['Date last evaluated'] = df_summary[
        'Date last evaluated'
    ].fillna(str(date.today()))

    # Catch if workbook has value for date last evaluated which is not datetime
    # compatible
    # Can test with first item in series as all rows have the same date value
    try:
        r = bool(date_parser.parse(str(df_summary['Date last evaluated'][0])))
    except date_parser._parser.ParserError:
        error_msg = (
            f"Value for date last evaluated \"{date_evaluated}\" is not "
            "compatible with datetime conversion"
        )
        return df_summary, error_msg


    df_summary["Date last evaluated"] = pd.to_datetime(
        df_summary["Date last evaluated"]
    )
    df_summary["Institution"] = config_variable["info"]["Institution"]
    df_summary["Collection method"] = config_variable["info"][
        "Collection method"
    ]
    df_summary["Allele origin"] = config_variable["info"]["Allele origin"]
    df_summary["Affected status"] = config_variable["info"]["Affected status"]

    # getting the folder name of workbook
    # the folder name should return designated folder for either CUH or NUH
    folder_name = get_folder(filename)
    if folder_name == config_variable["info"]["CUH folder"]:
        df_summary["Organisation"] = config_variable["info"][
            "CUH Organisation"
        ]

        df_summary["OrganisationID"] = config_variable["info"]["CUH org ID"]
    elif folder_name == config_variable["info"]["NUH folder"]:
        df_summary["Organisation"] = config_variable["info"][
            "NUH Organisation"
        ]

        df_summary["OrganisationID"] = config_variable["info"]["NUH org ID"]
    else:
        print("Running for the wrong folder")
        sys.exit(1)

    return df_summary

def get_included_fields(filename: str) -> pd.DataFrame:
    """
    Extract data from included sheet of variant workbook

    Parameters
    ----------
      variant workbook file name

    Return
    ------
      data frame from included sheet
    """
    workbook = load_workbook(filename)
    num_variants = workbook["summary"]["C38"].value
    interpreted_col = get_col_letter(workbook["included"], "Interpreted")
    df = pd.read_excel(
        filename,
        sheet_name="included",
        usecols=f"A:{interpreted_col}",
        nrows=num_variants,
    )
    df_included = df[
        [
            "CHROM",
            "POS",
            "REF",
            "ALT",
            "SYMBOL",
            "HGVSc",
            "Consequence",
            "Interpreted",
            "Comment",
        ]
    ].copy()
    if len(df_included["Interpreted"].value_counts()) > 0:
        df_included["Interpreted"] = df_included["Interpreted"].str.lower()
    df_included.rename(
        columns={
            "CHROM": "Chromosome",
            "SYMBOL": "Gene symbol",
            "POS": "Start",
            "REF": "Reference allele",
            "ALT": "Alternate allele",
        },
        inplace=True,
    )
    df_included["Local ID"] = ""
    for row in range(df_included.shape[0]):
        unique_id = uuid.uuid1()
        df_included.loc[row, "Local ID"] = f"uid_{unique_id.time}"
        time.sleep(0.5)
    df_included["Linking ID"] = df_included["Local ID"]

    return df_included

def get_report_fields(
    filename: str, df_included: pd.DataFrame
):  # -> tuple[pd.DataFrame, str]
    """
    Extract data from interpret sheet(s) of variant workbook

    Parameters
    ----------
      variant workbook file name
      data frame from included sheet

    Return
    ------
      data frame from interpret sheet(s)
      str for error message

    """
    workbook = load_workbook(filename)
    field_cells = [
        ("Associated disease", "C4"),
        ("Known inheritance", "C5"),
        ("Prevalence", "C6"),
        ("HGVSc", "C3"),
        ("Germline classification", "C26"),
        ("PVS1", "H10"),
        ("PVS1_evidence", "C10"),
        ("PS1", "H11"),
        ("PS1_evidence", "C11"),
        ("PS2", "H12"),
        ("PS2_evidence", "C12"),
        ("PS3", "H13"),
        ("PS3_evidence", "C13"),
        ("PS4", "H14"),
        ("PS4_evidence", "C14"),
        ("PM1", "H15"),
        ("PM1_evidence", "C15"),
        ("PM2", "H16"),
        ("PM2_evidence", "C16"),
        ("PM3", "H17"),
        ("PM3_evidence", "C17"),
        ("PM4", "H18"),
        ("PM4_evidence", "C18"),
        ("PM5", "H19"),
        ("PM5_evidence", "C19"),
        ("PM6", "H20"),
        ("PM6_evidence", "C20"),
        ("PP1", "H21"),
        ("PP1_evidence", "C21"),
        ("PP2", "H22"),
        ("PP2_evidence", "C22"),
        ("PP3", "H23"),
        ("PP3_evidence", "C23"),
        ("PP4", "H24"),
        ("PP4_evidence", "C24"),
        ("BS1", "K16"),
        ("BS1_evidence", "C16"),
        ("BS2", "K12"),
        ("BS2_evidence", "C12"),
        ("BS3", "K13"),
        ("BS3_evidence", "C13"),
        ("BA1", "K9"),
        ("BA1_evidence", "C9"),
        ("BP2", "K17"),
        ("BP2_evidence", "C17"),
        ("BP3", "K18"),
        ("BP3_evidence", "C18"),
        ("BS4", "K21"),
        ("BS4_evidence", "C21"),
        ("BP1", "K22"),
        ("BP1_evidence", "C22"),
        ("BP4", "K23"),
        ("BP4_evidence", "C23"),
        ("BP5", "K24"),
        ("BP5_evidence", "C24"),
        ("BP7", "K25"),
        ("BP7_evidence", "C25"),
    ]
    col_name = [i[0] for i in field_cells]
    df_report = pd.DataFrame(columns=col_name)
    report_sheets = [
        idx
        for idx in workbook.sheetnames
        if idx.lower().startswith("interpret")
    ]

    for idx, sheet in enumerate(report_sheets):
        for field, cell in field_cells:
            if workbook[sheet][cell].value is not None:
                df_report.loc[idx, field] = workbook[sheet][cell].value
    df_report.reset_index(drop=True, inplace=True)
    error_msg = None
    if not df_report.empty:
        error_msg = check_interpret_table(df_report, df_included)
    if not error_msg:
        # put strength as nan if it is 'NA'
        for row in range(df_report.shape[0]):
            for column in range(5, df_report.shape[1], 2):
                if df_report.iloc[row, column] == "NA":
                    df_report.iloc[row, column] = np.nan

        # removing evidence value if no strength
        for row in range(df_report.shape[0]):
            for column in range(5, df_report.shape[1], 2):
                if df_report.isnull().iloc[row, column]:
                    df_report.iloc[row, column + 1] = np.nan

        # getting comment on classification for clinvar submission
        matched_strength = [
            ("PVS", "Very Strong"),
            ("PS", "Strong"),
            ("PM", "Moderate"),
            ("PP", "Supporting"),
            ("BS", "Supporting"),
            ("BA", "Stand-Alone"),
            ("BP", "Supporting"),
        ]
        print(df_report)
        print(type(df_report))
        df_report["Comment on classification"] = ""
        for row in range(df_report.shape[0]):
            evidence = []
            for column in range(5, df_report.shape[1] - 1, 2):
                if not df_report.isnull().iloc[row, column]:
                    evidence.append(
                        [
                            df_report.columns[column],
                            df_report.iloc[row, column],
                        ]
                    )
            for index, value in enumerate(evidence):
                for st1, st2 in matched_strength:
                    if st1 in value[0] and st2 == value[1]:
                        evidence[index][1] = ""
            evidence_pair = []
            for e in evidence:
                evidence_pair.append("_".join(e).rstrip("_"))
            comment_on_classification = ",".join(evidence_pair)
            df_report.iloc[
                row, df_report.columns.get_loc("Comment on classification")
            ] = comment_on_classification

    return df_report


def determine_assembly(ref_genome):
    '''
    Work out assembly from the reference genome used to by VEP to process the
    data
    In our dias pipeline, this is the RefSeq cache in VEP 105

    For GRCh37, this will be GRCh37.p13
    For GRCh38, this will be GRCh38.p13

    Inputs:
        ref_genome (str): name of the reference genome for VEP for annotation
    Outputs:
        assembly (str): genome build of the reference genome (GRCh37 or GRCh38)
    '''
    if ref_genome == "GRCh37.p13":
        assembly = "GRCh37"
        print(
            f"Selected GRCh37 as assembly, because ref genome is {ref_genome}"
        )
    elif ref_genome == "GRCh38.p13":
        assembly = "GRCh38"
        print(
            f"Selected GRCh38 as assembly, because ref genome is {ref_genome}"
        )
    else:
        raise RuntimeError(
            f"Could not determine genome build from ref genome {ref_genome}"
        )
    return assembly


def add_lab_specific_guidelines(organisation_id, clinvar_dict):
    '''
    Format submission correctly if this is an NUH case
    Inputs:
        organisation_id (int): ClinVar organisation ID for submitting lab (CUH
        or NUH)
        clinvar_dict (dict): dictionary of info to submit to clinvar
    Outputs:
        clinvar_dict (dict): dictionary of info to submit to clinvar, edited
        to add url for assertion criteria which is specific to CUH or NUH
    '''
    # If NUH
    if organisation_id == 509428:
        clinvar_dict['assertionCriteria'] = {'url': 'https://submit.ncbi.nlm.n'
        'ih.gov/api/2.0/files/iptxgqju/uk-practice-guidelines-for-variant-clas'
        'sification-v4-01-2020.pdf/?format=attachment'}

    # If CUH
    elif organisation_id == 288359:
        clinvar_dict['assertionCriteria'] = {'url': 'https://submit.ncbi.nlm.n'
        'ih.gov/api/2.0/files/kf4l0sn8/uk-practice-guidelines-for-variant-clas'
        'sification-v4-01-2020.pdf/?format=attachment'}

    else:
        raise ValueError(
            f"Value given for organisation ID {organisation_id} is not a valid"
            " option.\nValid options:\n288359 - CUH\n509428 - NUH"
        )
    return clinvar_dict


def select_api_url(testing):
    '''
    Select which API URL to use depending on if this is a test run or if
    variants are planned to be submitted to ClinVar
    '''
    if testing in ["True", True, "true"]:
        api_url = "https://submit.ncbi.nlm.nih.gov/apitest/v1/submissions"
        print(
            f"Running in test mode, using {api_url}"
        )
    elif testing in ["False", False, "false"]:
        api_url = "https://submit.ncbi.nlm.nih.gov/api/v1/submissions/"
        print(
            f"Running in live mode, using {api_url}"
        )
    else:
        raise RuntimeError(
            f"Value for testing {testing} neither True nor False. Please "
            "specify whether this is a test run or not."
        )
    return api_url

def get_col_letter(worksheet: object, col_name: str) -> str:
    """
    Getting the column letter with specific col name

    Parameters
    ----------
    openpyxl object of current sheet
    str for name of column to get col letter

    Return
    -------
        str for column letter for specific column name
    """
    col_letter = None
    for column_cell in worksheet.iter_cols(1, worksheet.max_column):
        if column_cell[0].value == col_name:
            col_letter = column_cell[0].column_letter

    return col_letter

def check_interpret_table(
    df_report: pd.DataFrame, df_included: pd.DataFrame
) -> str:
    """
    check if ACMG classification and HGVSc are correctly
    filled in in the interpret table(s)

    Parameters
    ----------
      df from interpret sheet(s)
      df from included sheet

    Return
    ------
      str for error message
    """
    error_msg = []
    strength_dropdown = [
        "Very Strong",
        "Strong",
        "Moderate",
        "Supporting",
        "NA",
    ]
    BA1_dropdown = [
        "Stand-Alone",
        "Very Strong",
        "Strong",
        "Moderate",
        "Supporting",
        "NA",
    ]
    for row in range(df_report.shape[0]):
        try:
            assert (
                df_report.loc[row, "Germline classification"] is not np.nan
            ), "empty ACMG classification in interpret table"
            assert df_report.loc[row, "Germline classification"] in [
                "Pathogenic",
                "Likely Pathogenic",
                "Uncertain Significance",
                "Likely Benign",
                "Benign",
            ], "wrong ACMG classification in interpret table"
            assert (
                df_report.loc[row, "HGVSc"] is not np.nan
            ), "empty HGVSc in interpret table"
            assert df_report.loc[row, "HGVSc"] in list(df_included["HGVSc"]), (
                "HGVSc in interpret table does not match with that in "
                "included sheet"
            )
            criteria_list = [
                "PVS1",
                "PS1",
                "PS2",
                "PS3",
                "PS4",
                "PM1",
                "PM2",
                "PM3",
                "PM4",
                "PM5",
                "PM6",
                "PP1",
                "PP2",
                "PP3",
                "PP4",
                "BS2",
                "BS3",
                "BS1",
                "BP2",
                "BP3",
                "BS4",
                "BP1",
                "BP4",
                "BP5",
                "BP7",
            ]
            for criteria in criteria_list:
                if df_report.loc[row, criteria] is not np.nan:
                    assert (
                        df_report.loc[row, criteria] in strength_dropdown
                    ), f"Wrong strength in {criteria}"

            if df_report.loc[row, "BA1"] is not np.nan:
                assert (
                    df_report.loc[row, "BA1"] in BA1_dropdown
                ), "Wrong strength in BA1"

        except AssertionError as msg:
            error_msg.append(str(msg))
            print(msg)
    error_msg = "".join(error_msg)

    return error_msg

def checking_sheets(filename: str) -> str:
    """
    check if extra row(s)/col(s) are added in the sheets

    Parameters
    ----------
      variant workbook file name

    Return
    ------
      str for error message
    """
    workbook = load_workbook(filename)
    summary = workbook["summary"]
    reports = [
        idx
        for idx in workbook.sheetnames
        if idx.lower().startswith("interpret")
    ]
    try:
        assert (
            summary["G21"].value == "Date"
        ), "extra col(s) added or change(s) done in summary sheet"
        for sheet in reports:
            report = workbook[sheet]
            assert report["B26"].value == "FINAL ACMG CLASSIFICATION", (
                "extra row(s) or col(s) added or change(s) done in "
                "interpret sheet"
            )
            assert report["L8"].value == "B_POINTS", (
                "extra row(s) or col(s) added or change(s) done in "
                "interpret sheet"
            )
        error_msg = None
    except AssertionError as msg:
        error_msg = str(msg)
        print(msg)

    return error_msg


def check_interpreted_col(df: pd.DataFrame) -> str:
    """
    check if interpreted col in included sheet
    is correctly filled in

    Parameters
    ----------
    merged df

    Return
    ------
      str for error message
    """
    row_yes = df[df["Interpreted"] == "yes"].index.tolist()
    error_msg = []
    for row in range(df.shape[0]):
        if row in row_yes:
            try:
                assert (
                    df.loc[row, "Germline classification"] is not np.nan
                ), f"Wrong interpreted column in row {row+1} of included sheet"
            except AssertionError as msg:
                error_msg.append(str(msg))
                print(msg)
        else:
            try:
                assert df.loc[row, "Interpreted"] == "no", (
                    f"Wrong interpreted column dropdown in row {row+1} "
                    "of included sheet"
                )
                assert (
                    df.loc[row, "Germline classification"] is np.nan
                ), f"Wrong interpreted column in row {row+1} of included sheet"
            except AssertionError as msg:
                error_msg.append(str(msg))
                print(msg)
    error_msg = " ".join(error_msg)

    return error_msg


def check_sample_name(
    instrumentID: str,
    sample_ID: str,
    batchID: str,
    testcode: str,
    probesetID: str,
) -> str:
    """
    checking if individual parts of sample name have
    expected naming format

    Parameters
    ----------
      str values for instrumentID, sample_ID, batchID, testcode,
      probesetID

    Return
    ------
      str for error message
    """
    try:
        assert re.match(
            r"^\d{9}$", instrumentID
        ), "Unusual name for instrumentID"
        assert re.match(r"^\d{5}[A-Z]\d{4}$", sample_ID), "Unusual sampleID"
        assert re.match(r"^\d{2}[A-Z]{5}\d{1,}$", batchID), "Unusual batchID"
        assert re.match(r"^\d{4}$", testcode), "Unusual testcode"
        assert 0 < len(probesetID) < 20, "probesetID is too long/short"
        assert (
            probesetID.isalnum() and not probesetID.isalpha()
        ), "Unusual probesetID"
        error_msg = None
    except AssertionError as msg:
        error_msg = str(msg)
        print(msg)

    query = (
        ""
    )


    """
    check if extra row(s)/col(s) are added in the sheets

    Parameters
    ----------
      variant workbook file name

    Return
    ------
      str for error message
    """
    workbook = load_workbook(filename)
    summary = workbook["summary"]
    reports = [
        idx
        for idx in workbook.sheetnames
        if idx.lower().startswith("interpret")
    ]
    try:
        assert (
            summary["G21"].value == "Date"
        ), "extra col(s) added or change(s) done in summary sheet"
        for sheet in reports:
            report = workbook[sheet]
            assert report["B26"].value == "FINAL ACMG CLASSIFICATION", (
                "extra row(s) or col(s) added or change(s) done in "
                "interpret sheet"
            )
            assert report["L8"].value == "B_POINTS", (
                "extra row(s) or col(s) added or change(s) done in "
                "interpret sheet"
            )
        error_msg = None
    except AssertionError as msg:
        error_msg = str(msg)
        print(msg)

        query = (
        ""
    )

def submission_status_check(submission_id, headers, api_url):
    '''
    Queries ClinVar API about a submission ID to obtain more details about its
    submission record.
    Inputs:
        submission_id:  the generated submission id from ClinVar when a
        submission has been posted to their API
        headers: the required API url
    Outputs:
        status_response: the API response
    '''

    url = os.path.join(api_url, submission_id, "actions")
    response = requests.get(url, headers=headers)
    response_content = response.content.decode("UTF-8")
    for k, v in response.headers.items():
        print(f"{k}: {v}")
    print(response_content)
    if response.status_code not in [200]:
        raise RuntimeError(
            "Status check failed:\n" + str(headers) + "\n" + url
            + "\n" + response_content
        )

    status_response = json.loads(response_content)

    # Load summary file
    action = status_response["actions"][0]
    status = action["status"]
    print(f"Submission {submission_id} has status {status}")

    responses = action["responses"]
    if len(responses) == 0:
        print("Status 'responses' field had no items, check back later")
    else:
        print(
            "Status response had a response, attempting to "
            "retrieve any files listed"
        )
        try:
            f_url = responses[0]["files"][0]["url"]
        except (KeyError, IndexError) as error:
            f_url = None
            print(
                f"Error retrieving files: {error}.\n No API url for summary"
                "file found. Cannot query API for summary file based on "
                f"response {responses}"
            )

        if f_url is not None:
            print("GET " + f_url)
            f_response = requests.get(f_url, headers=headers)
            f_response_content = f_response.content.decode("UTF-8")
            if f_response.status_code not in [200]:
                raise RuntimeError(
                    "Status check summary file fetch failed:"
                    f"{f_response_content}"
                )
            file_content = json.loads(f_response_content)
            status_response = file_content

    return status_response