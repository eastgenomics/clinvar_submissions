import glob
import os
import json
from utils import utils
import pandas as pd
from pathlib import Path
import unittest
from openpyxl import load_workbook
from freezegun import freeze_time
import numpy as np

TEST_DATA_DIR = f"{Path(__file__).parent.resolve()}/test_data"

nuh_workbooks = glob.glob(TEST_DATA_DIR + '/NUH/' + "*.xlsx")
cuh_workbooks = glob.glob(TEST_DATA_DIR + '/CUH/' + "*.xlsx")

with open(TEST_DATA_DIR + '/test_config.json') as f:
    config = json.load(f)


for wb in nuh_workbooks + cuh_workbooks:
    filename = os.path.basename(wb).strip(".xlsx")
    locals()[filename] = wb


class TestParsing(unittest.TestCase):
    cuh_workbook = load_workbook(cuh)
    nuh_workbook = load_workbook(nuh)
    df_included = utils.get_included_fields(cuh_workbook, cuh)
    df_report, msg = utils.get_report_fields(cuh_workbook, config, df_included)

    def test_get_folder(self):
        """
        Test "get_folder" generates the correct folder
        where the workbook exists
        """
        with self.subTest("NUH folder correct"):
            NUH_folder = utils.get_folder_of_input_file(nuh)
            assert NUH_folder == "NUH"

        with self.subTest("CUH folder correct"):
            NUH_folder = utils.get_folder_of_input_file(cuh)
            assert NUH_folder == "CUH"

    def test_get_included_fields(self):
        """
        Test "get_included_fields" generates df with expected shape
        (2 rows and 11 columns in test case)
        Test generated df has expected columns
        Test generated df has expected interpreted column values
        in lowercase (no,yes in test case)
        Test contents of df and see if they are as expected
        """
        df = self.df_included

        with self.subTest("df has two variants, so two rows"):
            assert df.shape[0] == 2

        with self.subTest("df has expected number of columns"):
            assert df.shape[1] == 11

        with self.subTest("df has correct column headings"):
            assert list(df.columns) == [
                "chromosome",
                "start",
                "reference_allele",
                "alternate_allele",
                "gene_symbol",
                "hgvsc",
                "consequence",
                "classification_comment",
                "interpreted",
                "local_id",
                "linking_id",
            ]

        with self.subTest("one variant is interpreted, one is not"):
            assert list(df["interpreted"]) == ["no", "yes"]

        with self.subTest("chromosome extracted correctly"):
            assert list(df["chromosome"]) == [9, 16]

        with self.subTest("start positions extracted correctly"):
            assert list(df["start"]) == [135773000, 2134478]

        with self.subTest("ref allele extracted correctly"):
            assert list(df["reference_allele"]) == ["G", "C"]

        with self.subTest("alt allele extracted correctly"):
            assert list(df["alternate_allele"]) == ["GAA", "T"]

        with self.subTest("gene symbol extracted correctly"):
            assert list(df["gene_symbol"]) == ["TSC1", "TSC2"]

        with self.subTest("consequence extracted correctly"):
            assert list(df["consequence"]) == [
                "intron_variant&splice_polypyrimidine_tract_variant&splice_region_variant",
                "stop_gained"
            ]

        with self.subTest("HGVSc extracted correctly"):
            assert list(df["hgvsc"]) == [
                "NM_000368.5:c.2626-5_2626-4dup", "NM_000548.5:c.4255C>T"
            ]

    def test_get_report_fields(self):
        """
        Test "get_report_fields" generates df with expected shape
        (1 row and 58 columns in test case)
        Test generated df has expected columns
        Test generated df has the correct values for "HGVSc" and
        "Germline classification"
        Test generated df has the correct values for the strength
        and evidence columns (some are expected to be np.nan)
        Test there is no error message
        """
        df = self.df_report

        with self.subTest("df has one reported variants, so one row"):
            assert df.shape[0] == 1

        with self.subTest("df has expected number of columns"):
            assert df.shape[1] == 58

        with self.subTest("df has correct column names"):
            assert list(df.columns) == [
                "associated_disease",
                "known_inheritance",
                "prevalence",
                "hgvsc",
                "germline_classification",
                'pvs1',
                'pvs1_evidence',
                'ps1',
                'ps1_evidence',
                'ps2',
                'ps2_evidence',
                'ps3',
                'ps3_evidence',
                'ps4',
                'ps4_evidence',
                'pm1',
                'pm1_evidence',
                'pm2',
                'pm2_evidence',
                'pm3',
                'pm3_evidence',
                'pm4',
                'pm4_evidence',
                'pm5',
                'pm5_evidence',
                'pm6',
                'pm6_evidence',
                'pp1',
                'pp1_evidence',
                'pp2',
                'pp2_evidence',
                'pp3',
                'pp3_evidence',
                'pp4',
                'pp4_evidence',
                'bs1',
                'bs1_evidence',
                'bs2',
                'bs2_evidence',
                'bs3',
                'bs3_evidence',
                'ba1',
                'ba1_evidence',
                'bp2',
                'bp2_evidence',
                'bp3',
                'bp3_evidence',
                'bs4',
                'bs4_evidence',
                'bp1',
                'bp1_evidence',
                'bp4',
                'bp4_evidence',
                'bp5',
                'bp5_evidence',
                'bp7',
                'bp7_evidence',
                "comment_on_classification",
            ]

        with self.subTest("df has correct HGVSc"):
            assert df["hgvsc"][0] == "NM_000548.5:c.4255C>T"

        with self.subTest("df has correct classification"):
            assert df["germline_classification"][0] == "Pathogenic"

        with self.subTest("df has correct comment on classification"):
            assert df["comment_on_classification"][0] == "PVS1,PS4_Moderate"

        with self.subTest("df correct has strength for PVS1"):
            assert df["pvs1"][0] == "Very Strong"

        with self.subTest("df correct has strength for PS4"):
            assert df["ps4"][0] == "Moderate"

        with self.subTest("pvs1 evidence extracted correctly"):
            assert df["pvs1_evidence"][0] == (
                    "Exon present in all transcripts on gnomAD. "
                    "LOF known mechanism of disease. "
                    "Predicted to undergo nonsense-mediated decay."
                )

        with self.subTest("ps4 evidence extracted correctly"):
            assert df["ps4_evidence"][0] == (
                "PMID: 10205261 (1 case, Roach et al criteria), "
                "35870981 (1 case, 2012 International TS Complex "
                "Consensus Conference criteria), 12111193 "
                "(1 case, Roach et al criteria), 28065512 "
                "(1 case, 2012 International TS Complex "
                "Consensus Conference criteria)."
            )

        with self.subTest("null criteria is null in df"):
            assert pd.isna(df["pp3"][0])

        with self.subTest("null criteria have null evidence in df"):
            assert pd.isna(df["pp3_evidence"][0])

        with self.subTest("'NA' criteria changed to null in df"):
            assert pd.isna(df["pp1"][0])

        with self.subTest(
            "null criteria have null evidence in df, even if evidence provided"
        ):
            assert pd.isna(df["pp1_evidence"][0])

        with self.subTest("No error message should be returned"):
            assert self.msg is None

    def test_check_interpret_table_returns_no_error_if_hgvsc_matches(self):
        error_msg = utils.check_interpret_table(
            self.df_report,
            self.df_included,
            config
        )
        assert error_msg is None

    def test_check_interpret_table_error_if_hgvsc_wrong(self):
        wrong_hgvsc_workbook = load_workbook(cuh_wrong_hgvsc)
        df_include = utils.get_included_fields(
            wrong_hgvsc_workbook, cuh_wrong_hgvsc
        )
        df_report, msg = utils.get_report_fields(
            wrong_hgvsc_workbook, config, df_include
        )
        error_msg = utils.check_interpret_table(
            df_report,
            df_include,
            config
        )
        assert error_msg == (
            "HGVSc in interpret table does not match with that in included "
            "sheet"
        )

    def test_check_interpret_table_error_if_hgvsc_empty(self):
        empty_hgvsc_workbook = load_workbook(cuh_empty_hgvsc)
        df_include = utils.get_included_fields(
            empty_hgvsc_workbook, cuh_empty_hgvsc
        )
        df_report, msg = utils.get_report_fields(
            empty_hgvsc_workbook, config, df_include
        )
        error_msg = utils.check_interpret_table(
            df_report,
            df_include,
            config
        )
        assert error_msg == "empty HGVSc in interpret table"

    def test_check_interpret_table_error_if_no_acmg_classification(self):
        no_acmg_workbook = load_workbook(cuh_empty_acmg)
        df_include = utils.get_included_fields(
            no_acmg_workbook, cuh_empty_acmg
        )
        df_report, msg = utils.get_report_fields(
            no_acmg_workbook, config, df_include
        )
        error_msg = utils.check_interpret_table(
            df_report,
            df_include,
            config
        )
        assert error_msg == "empty ACMG classification in interpret table"

    def test_check_interpret_table_error_if_wrong_acmg_classification(self):
        wrong_acmg_workbook = load_workbook(cuh_wrong_acmg)
        df_include = utils.get_included_fields(
            wrong_acmg_workbook, cuh_wrong_acmg
        )
        df_report, msg = utils.get_report_fields(
            wrong_acmg_workbook, config, df_include
        )
        error_msg = utils.check_interpret_table(
            df_report,
            df_include,
            config
        )
        assert error_msg == "wrong ACMG classification in interpret table"

    def test_check_interpret_table_error_if_wrong_strength(self):
        wrong_interpret_strength_workbook = load_workbook(
            nuh_wrong_interpret_strength
        )
        df_include = utils.get_included_fields(
            wrong_interpret_strength_workbook, nuh_wrong_interpret_strength
        )
        df_report, msg = utils.get_report_fields(
            wrong_interpret_strength_workbook, config, df_include
        )
        error_msg = utils.check_interpret_table(
            df_report,
            df_include,
            config
        )
        assert error_msg == "Wrong strength in pm2"

    def test_checking_sheet_wrong_summary(self):
        wrong_summary_workbook = load_workbook(nuh_wrong_summary)
        msg = utils.checking_sheets(wrong_summary_workbook)
        assert msg == "extra col(s) added or change(s) done in summary sheet"

    def test_checking_sheet_wrong_interpret_row(self):
        wrong_interpret_row = load_workbook(nuh_wrong_interpret_row)
        msg = utils.checking_sheets(wrong_interpret_row)
        assert msg == (
            "extra row(s) or col(s) added or change(s) done in interpret sheet"
        )

    def test_get_summary_fields(self):
        """
        Test "get_summary_fields" generates df with expected shape
        (1 row and 15 columns in test case)
        Test the "Preferred condition name" is split as expected
        (Tuberous sclerosis in test case)
        Test if the "Ref genome" is correctly extracted
        (GRCh37.p13 in test case)
        Test if the "Date last evaluated" is pd date time format
        """
        df, msg = utils.get_summary_fields(self.nuh_workbook, config, nuh)

        with self.subTest("df has expected number of rows"):
            assert df.shape[0] == 1

        with self.subTest("df has expected number of columns"):
            assert df.shape[1] == 16

        with self.subTest("correct preferred condition name extracted"):
            assert df["preferred_condition_name"][0] == (
                "Inherited breast cancer and ovarian cancer;Inherited breast "
                "cancer and ovarian cancer"
            )

        with self.subTest("Ref genome correctly extracted"):
            assert df["ref_genome"][0] == "GRCh37.p13"

        with self.subTest("Vale for date_last_evaluated is date type"):
            assert isinstance(
                df["date_last_evaluated"][0],
                pd._libs.tslibs.timestamps.Timestamp
            )

    @freeze_time("2024-07-10 22:22:22")
    def test_no_evaluated_date(self):
        '''
        Test that when there is no date last evaluated, today's date is used.
        Expect the time to change to 00:00 as we are only using date not time.
        This test uses an example workbook with nothing in the date cell
        '''
        no_evaluated_date_workbook = load_workbook(nuh_no_evaluated_date)
        df, msg = utils.get_summary_fields(
            no_evaluated_date_workbook, config, nuh_no_evaluated_date
        )
        assert str(df['date_last_evaluated'].item()) == "2024-07-10 00:00:00"

    def test_error_message_if_date_last_evaluated_invalid(self):
        '''
        Test that if the workbook has an evaluation date that is not compatible
        with datetime i.e. is not a date, the correct error message is returned
        which will cause this workbook to be skipped and added to parsing
        failed list.
        This test uses an example workbook with "Not valid" in the date cell
        '''
        invalid_evaluation_date_workbook = load_workbook(
            nuh_invalid_evaluated_date
        )
        df, msg = utils.get_summary_fields(
            invalid_evaluation_date_workbook,
            config,
            nuh_invalid_evaluated_date
        )
        assert msg == (
            "Value for date last evaluated \"Not valid\" is not compatible "
            "with datetime conversion"
        )

    def test_interpreted_col_correct(self):
        """
        Test if interpreted col (yes/no) is correctly filled in no error is
        raised.
        """
        df_summary, _ = utils.get_summary_fields(
            self.cuh_workbook, config, cuh
        )
        df_merged = pd.merge(self.df_included, df_summary, how="cross")
        df_final = pd.merge(df_merged, self.df_report, on="hgvsc", how="left")
        msg = utils.check_interpreted_col(df_final)
        assert msg is None

    def test_interpreted_col_errors_if_incongruous(self):
        '''
        Test that there is a classification for each variant that has been
        interpreted and no classification if not interpreted.
        '''
        data = [
            ['no', 'Pathogenic', 'NM_000368.5:c.2626-5_2626-4dup'],
            ['yes', np.nan, 'NM_000548.5:c.4255C>T']
        ]
        df = pd.DataFrame(
            data, columns=["interpreted", "germline_classification", "hgvsc"]
        )

        error_msg = utils.check_interpreted_col(df)
        assert error_msg == (
            "Variant NM_000548.5:c.4255C>T has interpreted = yes, but no final"
            " classification could be extracted from interpret sheets. Variant"
            " NM_000368.5:c.2626-5_2626-4dup has interpreted = no, but a final"
            " classification could be extracted from interpret sheets."
        )

    def test_interpreted_col_errors_if_value_neither_yes_or_no(self):
        data = [
            ['yes', 'Pathogenic', 'NM_000368.5:c.2626-5_2626-4dup'],
            ['foo', np.nan, 'NM_000548.5:c.4255C>T']
        ]
        df = pd.DataFrame(
            data, columns=["interpreted", "germline_classification", "hgvsc"]
        )

        error_msg = utils.check_interpreted_col(df)
        assert error_msg == (
            "Values in interpreted column are not all either 'yes' or 'no'"
        )
